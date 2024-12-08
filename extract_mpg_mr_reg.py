#!/common/appl/python/python-3.6.4/bin/python3
import warnings
import argparse
import openpyxl

#heree
f = open("demofile2.txt", "a")

warnings.filterwarnings("ignore", category=UserWarning, message="DrawingML support is incomplete.*")

def get_ba(wb, hier_value):
    sheet = wb["Register Configuration"]

    ba_str = None

    for row in range(3, 29):
        hier_cell = sheet[f"C{row}"].value
        if hier_cell == hier_value:
            ba_str = sheet[f"D{row}"].value
            break

    if ba_str is None:
        print(f"Error: No matching '{hier_value}' found in 'the Hier Structure' column of 'Register Configuration' sheet.")
        ba_value = 0
    else:
        hex_value = ba_str.split("'")[1].replace("_", "")
        ba_value = int(hex_value, 16)

        pkcprot0_address = hex(ba_value + 0x0CF0)
        pkcprot1_address = hex(ba_value + 0x0CF4)
        print(f"\n{hier_value}'s subsysss base address: 0x{hex_value.lower()}")
        print(f"{hier_value}'s MDLCnPKCPROT0 address: {pkcprot0_address}")
        print(f"{hier_value}'s MDLCnPKCPROT1 address: {pkcprot1_address}")

    return ba_value

def get_hier_sheet(hier_value):
    if 'VIP' in hier_value:
        hier_sheet = 'VIP'
    elif 'DDR' in hier_value:
        hier_sheet = 'DDR0to7'
    elif 'NPU' in hier_value:
        hier_sheet = 'NPU'
    elif 'IM' in hier_value:
        hier_sheet = 'IM'
    else:
        hier_sheet = hier_value
    return hier_sheet

def extract_MPG(wb, hier_value, mod_value, ba_value):

    hier_sheet = get_hier_sheet(hier_value)
    if hier_sheet not in wb.sheetnames:
        print(f"Error: Sheet '{hier_sheet}' does not exist")
        return
    
    sheet = wb[hier_sheet]
    pdid_value = None
    for row in range(6, sheet.max_row + 1):
        mod_cell = sheet[f"F{row}"].value
        if mod_cell == mod_value:
            pdid_value = sheet[f"AW{row}"].value
            break

    if pdid_value is None:
        print(f"Error: No matching MOD value '{mod_value}' found in sheet '{hier_sheet}'.")
    else:
        if pdid_value == '-':
            mpdg_address = 0
            mpdgs_address = 0
            print("\nInfo: PDID value is '-' (always ON power domain), no MPG setting.")
        else:
            pdid = int(pdid_value)
            mpdg_address = hex(ba_value + 0x0200 + pdid * 0x4)
            mpdgs_address = hex(ba_value + 0x0300 + pdid * 0x4)

            print(f"\n{mod_value}'s PDID (k): {pdid}")
            print(f"{mod_value}'s MDLCnMPDG  address: {mpdg_address}")
            print(f"{mod_value}'s MDLCnMPDGS address: {mpdgs_address}")

def extract_MR(wb, hier_value, mod_value, ba_value):

    hier_sheet = get_hier_sheet(hier_value)
    if hier_sheet not in wb.sheetnames:
        print(f"Error: Sheet '{hier_sheet}' does not exist")
        return None, None

    sheet = wb[hier_sheet]

    regid_value = None
    bitid_value = None

    for row in range(6, sheet.max_row + 1):
        mod_cell = sheet[f"F{row}"].value
        if mod_cell == mod_value:
            regid_value = sheet[f"J{row}"].value
            bitid_value = sheet[f"L{row}"].value
            break

    if regid_value is None:
        print(f"Error: No matching MOD value '{mod_value}' found in sheet '{hier_sheet}'.")
    else:
        if regid_value == '-':
            msress_address = 0
            msres_address = 0
            print("Error: REGID value is '-', MSRESS and MSRES addresses are set to 0.")
        else:
            regid = int(regid_value)

            msres_address = hex(ba_value + 0x0900 + regid * 0x4)
            msress_address = hex(ba_value + 0x0960 + regid * 0x4)

            print(f"\n{mod_value}'s REG No.: {regid}")
            print(f"{mod_value}'s BIT No.: {bitid_value}")
            print(f"{mod_value}'s MDLCnMSRES address: {msres_address}")
            print(f"{mod_value}'s MDLCnMSRESS address: {msress_address}")

            #heree

            f.write(f"{mod_value} {msress_address}")
            f.write(f"{mod_value} {msres_address}")



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Register for MPG and MR flow verification")
    parser.add_argument("-file", help="Path to the TS Excel file", required=True)
    parser.add_argument("-hier", help="HIER name", required=True)
    parser.add_argument("-mod", help="MOD name", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.file, data_only=True)
    subsysss_ba = get_ba(wb, args.hier)
    extract_MPG(wb, args.hier, args.mod, subsysss_ba)
    extract_MR (wb, args.hier, args.mod, subsysss_ba)

    wb.close()

f.close()
#heree
#open and read the file after the appending:
f = open("demofile2.txt", "r")
print(f.read())
